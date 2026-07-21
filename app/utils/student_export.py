from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_students_to_excel(students):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = [
        'Student ID',
        'Roll Number',
        'Full Name',
        'Email',
        'Phone',
        'Faculty',
        'Department',
        'Program',
        'Academic Year',
        'Semester',
        'Admission Year',
        'Verified',
        'Active'
    ]

    # Header style
    header_fill = PatternFill(
        start_color='1E40AF',
        end_color='1E40AF',
        fill_type='solid'
    )

    header_font = Font(
        color='FFFFFF',
        bold=True
    )

    # Write headers
    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_idx, student in enumerate(students, 2):

        ws.cell(row=row_idx, column=1, value=student.student_id)
        ws.cell(row=row_idx, column=2, value=student.roll_number)
        ws.cell(row=row_idx, column=3, value=student.full_name)
        ws.cell(row=row_idx, column=4, value=student.email)
        ws.cell(row=row_idx, column=5, value=student.phone)
        ws.cell(row=row_idx, column=6, value=student.faculty.name)
        ws.cell(row=row_idx, column=7, value=student.department.name)
        ws.cell(row=row_idx, column=8, value=student.program.name)
        ws.cell(row=row_idx, column=9, value=student.academic_year.name)
        ws.cell(row=row_idx, column=10, value=student.semester.name)
        ws.cell(row=row_idx, column=11, value=student.admission_year)
        ws.cell(row=row_idx, column=12, value='Yes' if student.is_verified else 'No')
        ws.cell(row=row_idx, column=13, value='Yes' if student.is_active else 'No')

    # Auto width
    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output